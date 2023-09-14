import time
import openpyxl
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from fake_useragent import UserAgent
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

common_waiting_seconds = 5


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    # print("xlsx格式表格写入数据成功！")


def append_excel_xlsx(path, sheet_name, new_data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    for row in new_data:
        # 在最后一行后面追加新数据
        sheet.append(row)
    workbook.save(path)


def crawl(headless=False, new_file=True, start=1000, end=98799):
    book_name_xlsx = 'result.xlsx'
    sheet_name_xlsx = 'sheet1'
    vv = [['邮编', '区域', '公司名', '地址', '电话']]
    if new_file:
        write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, vv)
    vv.clear()

    base_url = 'https://www.solistracteur.fr/Revendeurs-Solis'
    options = Options()
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument('user-agent=%s' % UserAgent().random)
    if headless:
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1080')

    driver = webdriver.Chrome(executable_path='chromedriver.exe', options=options)
    page_wait_time = 20
    driver.set_page_load_timeout(page_wait_time)

    for ii in range(start, end + 1):
        post_code = str(ii).zfill(5)
        try:
            driver.get(base_url)
        except TimeoutException as e:
            print("！！！！！！time out after %s seconds when loading page！！！！！！" % page_wait_time)
        finally:
            WebDriverWait(driver, common_waiting_seconds).until(
                lambda x: x.find_element(By.ID, 'annuaire_cp_commune_id'))
            print('======= 处理邮编 %s' % post_code)
            post_code_input = driver.find_element(By.ID, 'annuaire_cp_commune_id')
            post_code_input.send_keys(Keys.CONTROL, 'a')
            post_code_input.send_keys(post_code)
            # 手动出发 onChange 事件，目前不需要了
            # driver.execute_script("arguments[0].dispatchEvent(new Event('change'))", post_code_input)
            time.sleep(common_waiting_seconds)

            WebDriverWait(driver, common_waiting_seconds).until(lambda x: x.find_element(By.NAME, 'commune_id'))

            time.sleep(common_waiting_seconds)
            select_element = driver.find_element(By.XPATH,
                                                 '/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/form/span[1]/div/div[1]'
                                                 '/select')
            select = Select(select_element)
            n = len(select.options)
            if n == 0:
                print('此邮编 %s 无数据' % post_code)
                continue

            print('此邮编 %s 共有 %d 个区域' % (post_code, n))
            vv.clear()
            for i in range(n):
                time.sleep(common_waiting_seconds)
                driver.find_element(By.NAME, 'valider_rechecher').click()
                time.sleep(common_waiting_seconds)
                area = driver.find_element(By.XPATH,
                                           '/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/form/span[1]/div/div[1]'
                                           '/select') \
                    .find_element(By.CSS_SELECTOR, 'option[selected="selected"]').text
                print('开始处理第 %d 个区域: %s' % (i + 1, area))
                # time.sleep(5)
                info_list = driver.find_elements(By.XPATH,
                                                 "/html/body/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/div[1]"
                                                 "/div[@class='liste_revendeurs_un']")
                for info in info_list:
                    company_name = info.find_element(By.CLASS_NAME, 'liste_revendeurs_un_titre').text
                    company_address = info.find_element(By.XPATH, './/div[3]').text.split('\nTéléphone')[0]
                    company_phone_num = info.find_element(By.XPATH, './/a').text
                    vv.append([post_code, area, company_name, company_address, company_phone_num])
                down_clk = driver.find_element(By.XPATH,
                                               '/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/form/span[1]/div/div[1]'
                                               '/select')
                down_clk.click()
                down_clk.send_keys(Keys.DOWN)
                down_clk.send_keys(Keys.ENTER)
            # write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, vv)
            append_excel_xlsx(book_name_xlsx, sheet_name_xlsx, vv)
            vv.clear()


if __name__ == '__main__':
    crawl(headless=False, new_file=True, start=1000, end=98799)
